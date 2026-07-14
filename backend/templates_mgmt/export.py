import csv
import io
import json

from cycles.models import TaskDependency

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


# Export
#
# All three formats (JSON, CSV, XLSX) are written from the same
# intermediate shape, a plain dict, so there is exactly one place that
# turns a Template into that shape (template_to_intermediate). The
# formats only differ in how that dict gets written out, never in what
# a template's data actually is. Same "one canonical implementation"
# approach as scheduling.py and dependency_engine.py.
#
# local_id exists so a downloaded file's cross-references (a task's
# activity, a dependency's two ends, a tag assignment) are
# self-contained and readable on their own, rather than leaking
# internal database primary keys that mean nothing outside this
# system. Every reference column (activity, depends-on, tags) is
# ALWAYS paired with a resolved, human-readable name column right
# next to it — the local_id is there so the relationship is
# unambiguous even if two tasks share a name, the name is there so a
# person reading the file doesn't have to go cross-referencing IDs by
# hand. Repeating the name next to its ID is deliberate redundancy,
# not a mistake.

SUPPORTED_FORMATS = ("json", "csv", "xlsx")


def template_to_intermediate(template):
    """Template -> the canonical dict every export format is built from."""
    activities = list(template.template_activities.all())
    tasks = list(template.template_tasks.all())
    dependencies = list(
        TaskDependency.objects.filter(task__template=template).select_related(
            "task", "depends_on_task"
        )
    )

    activity_local_id = {a.template_activity_id: f"A{a.template_activity_id}" for a in activities}
    task_local_id = {t.template_task_id: f"T{t.template_task_id}" for t in tasks}
    activity_name_by_id = {a.template_activity_id: a.activity_name for a in activities}
    task_name_by_id = {t.template_task_id: t.task_name for t in tasks}

    # Tags: many-to-many, so each task/activity gets a plain list of
    # tag names. Deliberately just names (not local ids) — a tag has
    # no other data worth cross-referencing, unlike activities/deps.
    from .models import TemplateTaskTag, TemplateActivityTag

    task_tags = {}
    for tt in TemplateTaskTag.objects.filter(template_task__template=template).select_related("tag"):
        task_tags.setdefault(tt.template_task_id, []).append(tt.tag.tag_name)

    activity_tags = {}
    for at in TemplateActivityTag.objects.filter(template_activity__template=template).select_related("tag"):
        activity_tags.setdefault(at.template_activity_id, []).append(at.tag.tag_name)

    all_tags = sorted({name for names in task_tags.values() for name in names}
                       | {name for names in activity_tags.values() for name in names})

    return {
        "template": {
            "template_name": template.template_name,
            "description": template.description or "",
            "is_public": template.is_public,
            "category": template.category.category_name if template.category_id else "",
        },
        "activities": [
            {
                "local_id": activity_local_id[a.template_activity_id],
                "activity_name": a.activity_name,
                "description": a.description or "",
                "start_offset_days": a.start_offset_days,
                "end_offset_days": a.end_offset_days,
                "note_text": a.note_text or "",
                "tags": sorted(activity_tags.get(a.template_activity_id, [])),
            }
            for a in activities
        ],
        "tasks": [
            {
                "local_id": task_local_id[t.template_task_id],
                "task_name": t.task_name,
                "description": t.description or "",
                "day_offset": t.day_offset,
                "duration_days": t.duration_days,
                "is_mandatory": t.is_mandatory,
                "is_fixed_date": t.is_fixed_date,
                "reminder_lead_days": t.reminder_lead_days or [],
                "note_text": t.note_text or "",
                "activity_local_id": activity_local_id.get(t.template_activity_id, ""),
                "activity_name": activity_name_by_id.get(t.template_activity_id, ""),
                "tags": sorted(task_tags.get(t.template_task_id, [])),
            }
            for t in tasks
        ],
        "dependencies": [
            {
                "task_local_id": task_local_id[d.task_id],
                "task_name": task_name_by_id[d.task_id],
                "depends_on_local_id": task_local_id[d.depends_on_task_id],
                "depends_on_name": task_name_by_id[d.depends_on_task_id],
            }
            for d in dependencies
        ],
        "tags": all_tags,
    }


def write_json(data):
    return json.dumps(data, indent=2).encode("utf-8")


def _yn(value):
    return "Yes" if value else ""


def _tags_str(tags):
    return "; ".join(tags)


def write_csv(data):
    # CSV can only ever be one flat table, so this stays a
    # row_type-per-line format — but every column a human would
    # actually want (tags, resolved activity/dependency names,
    # friendly Yes/No booleans) is included, instead of leaving them
    # for a separate file the CSV format can't provide anyway.
    buffer = io.StringIO()
    fieldnames = [
        "row_type", "local_id", "name", "description",
        "day_offset", "duration_days", "start_offset_days", "end_offset_days",
        "mandatory", "fixed_date", "reminder_lead_days", "note_text",
        "activity_local_id", "activity_name",
        "depends_on_local_id", "depends_on_name",
        "tags", "is_public", "category",
    ]
    writer = csv.DictWriter(buffer, fieldnames=fieldnames)
    writer.writeheader()

    writer.writerow({
        "row_type": "template",
        "name": data["template"]["template_name"],
        "description": data["template"]["description"],
        "is_public": _yn(data["template"]["is_public"]),
        "category": data["template"]["category"],
    })

    for activity in data["activities"]:
        writer.writerow({
            "row_type": "activity",
            "local_id": activity["local_id"],
            "name": activity["activity_name"],
            "description": activity["description"],
            "start_offset_days": activity["start_offset_days"],
            "end_offset_days": activity["end_offset_days"],
            "note_text": activity["note_text"],
            "tags": _tags_str(activity["tags"]),
        })

    for task in data["tasks"]:
        writer.writerow({
            "row_type": "task",
            "local_id": task["local_id"],
            "name": task["task_name"],
            "description": task["description"],
            "day_offset": task["day_offset"],
            "duration_days": task["duration_days"] if task["duration_days"] is not None else "",
            "mandatory": _yn(task["is_mandatory"]),
            "fixed_date": _yn(task["is_fixed_date"]),
            "reminder_lead_days": ";".join(str(d) for d in task["reminder_lead_days"]),
            "note_text": task["note_text"],
            "activity_local_id": task["activity_local_id"],
            "activity_name": task["activity_name"],
            "tags": _tags_str(task["tags"]),
        })

    for dependency in data["dependencies"]:
        writer.writerow({
            "row_type": "dependency",
            "local_id": dependency["task_local_id"],
            "name": dependency["task_name"],
            "depends_on_local_id": dependency["depends_on_local_id"],
            "depends_on_name": dependency["depends_on_name"],
        })

    if data["tags"]:
        writer.writerow({"row_type": "tag_reference", "name": "; ".join(data["tags"])})

    return buffer.getvalue().encode("utf-8")


# ── XLSX styling ──────────────────────────────────────────────
_HEADER_FILL = None
_HEADER_FONT = None
_THIN_BORDER = None
_ZEBRA_FILL = None


def _init_styles():
    global _HEADER_FILL, _HEADER_FONT, _THIN_BORDER, _ZEBRA_FILL
    _HEADER_FILL = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    _HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="E2E8F0")
    _THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    _ZEBRA_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")


def _write_header_row(sheet, headers, widths):
    sheet.append(headers)
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 20


def _style_data_rows(sheet, start_row, end_row, n_cols):
    for row in range(start_row, end_row + 1):
        fill = _ZEBRA_FILL if (row - start_row) % 2 == 1 else None
        for col in range(1, n_cols + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = _THIN_BORDER
            if fill:
                cell.fill = fill


def write_xlsx(data):
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed, xlsx export is unavailable.")

    _init_styles()
    workbook = openpyxl.Workbook()

    # ── TEMPLATE (single record, shown as Field | Value — reads much
    # better than a one-row table for a record with only 4 fields) ──
    template_sheet = workbook.active
    template_sheet.title = "Template"
    _write_header_row(template_sheet, ["Field", "Value"], [22, 50])
    template_fields = [
        ("Template name", data["template"]["template_name"]),
        ("Description", data["template"]["description"]),
        ("Public", _yn(data["template"]["is_public"])),
        ("Category", data["template"]["category"]),
        ("Tasks", len(data["tasks"])),
        ("Activities", len(data["activities"])),
        ("Dependencies", len(data["dependencies"])),
    ]
    for label, value in template_fields:
        template_sheet.append([label, value])
    _style_data_rows(template_sheet, 2, 1 + len(template_fields), 2)
    for row in range(2, 2 + len(template_fields)):
        template_sheet.cell(row=row, column=1).font = Font(bold=True)

    # ── ACTIVITIES ──
    activities_sheet = workbook.create_sheet("Activities")
    headers = ["Local ID", "Name", "Description", "Start day", "End day", "Note", "Tags"]
    widths = [10, 24, 34, 10, 10, 30, 26]
    _write_header_row(activities_sheet, headers, widths)
    for activity in data["activities"]:
        activities_sheet.append([
            activity["local_id"], activity["activity_name"], activity["description"],
            activity["start_offset_days"], activity["end_offset_days"], activity["note_text"],
            _tags_str(activity["tags"]),
        ])
    _style_data_rows(activities_sheet, 2, 1 + len(data["activities"]), len(headers))

    # ── TASKS ──
    tasks_sheet = workbook.create_sheet("Tasks")
    headers = [
        "Local ID", "Name", "Description", "Day offset", "Duration (days)",
        "Mandatory", "Fixed date", "Reminders (days before)", "Note",
        "Activity ID", "Activity", "Tags",
    ]
    widths = [10, 22, 30, 11, 15, 11, 11, 20, 26, 11, 20, 26]
    _write_header_row(tasks_sheet, headers, widths)
    for task in data["tasks"]:
        tasks_sheet.append([
            task["local_id"], task["task_name"], task["description"],
            task["day_offset"], task["duration_days"],
            _yn(task["is_mandatory"]), _yn(task["is_fixed_date"]),
            ";".join(str(d) for d in task["reminder_lead_days"]),
            task["note_text"], task["activity_local_id"], task["activity_name"],
            _tags_str(task["tags"]),
        ])
    _style_data_rows(tasks_sheet, 2, 1 + len(data["tasks"]), len(headers))

    # ── DEPENDENCIES (resolved names alongside the local-id join) ──
    dependencies_sheet = workbook.create_sheet("Dependencies")
    headers = ["Task ID", "Task", "Depends On ID", "Depends On"]
    widths = [10, 24, 15, 24]
    _write_header_row(dependencies_sheet, headers, widths)
    for dependency in data["dependencies"]:
        dependencies_sheet.append([
            dependency["task_local_id"], dependency["task_name"],
            dependency["depends_on_local_id"], dependency["depends_on_name"],
        ])
    _style_data_rows(dependencies_sheet, 2, 1 + len(data["dependencies"]), len(headers))

    # ── TAGS (reference sheet — every distinct tag name used anywhere
    # in this template, so it's discoverable without scanning every
    # task/activity row by hand) ──
    tags_sheet = workbook.create_sheet("Tags")
    _write_header_row(tags_sheet, ["Tag name"], [26])
    for tag_name in data["tags"]:
        tags_sheet.append([tag_name])
    _style_data_rows(tags_sheet, 2, 1 + len(data["tags"]), 1)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


WRITERS = {"json": write_json, "csv": write_csv, "xlsx": write_xlsx}